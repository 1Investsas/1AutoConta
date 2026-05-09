"""
Exportador SIIGO — Fase 3.

Genera uno o varios archivos Excel (.xlsx) en el formato exacto del
template oficial de SIIGO Nube ("Modelo de importacion de comprobantes
contables.xlsx") para la importación de comprobantes contables.

Características:
  - Hoja "Datos", encabezados en fila 1 (rojo = obligatorio, azul = opcional).
  - Anchos de columna idénticos al template SIIGO.
  - Respeta el límite de 500 filas de SIIGO: si hay más filas genera
    múltiples archivos numerados (_parte1, _parte2, …).
  - Las filas con cuenta [PENDIENTE] se resaltan en rojo para que el
    contador las complete antes de subir el archivo.
  - Columnas numéricas (Débito / Crédito) sin fórmulas, solo valores.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import PreasientoContable
from app.config import SIIGO_MAX_FILAS_POR_ARCHIVO, OUTPUT_DIR
from app.siigo.mapeador import (
    ENCABEZADOS_SIIGO,
    _COLS_REQUERIDAS,
    FilaSiigo,
    mapear_lote,
    partir_en_chunks,
)

logger = logging.getLogger(__name__)

# Colores del template SIIGO
_COLOR_ROJO_REQUERIDO  = "FF0000"   # columnas obligatorias (1,2,3,6,7)
_COLOR_AZUL_OPCIONAL   = "0070C0"   # columnas opcionales

# Colores para filas con datos anómalos
_COLOR_PENDIENTE   = "FFE0E0"   # Rojo claro — cuenta pendiente
_COLOR_ADVERTENCIA = "FFF3CD"   # Amarillo — tercero no encontrado

# Anchos de columna tomados del template SIIGO (en unidades de carácter Excel)
_ANCHOS_TEMPLATE = [
    13.109375,   # A  — Tipo de comprobante
    13.5546875,  # B  — Consecutivo comprobante
    11.44140625, # C  — Fecha de elaboración
    8.6640625,   # D  — Sigla moneda
    14.5546875,  # E  — Tasa de cambio
    17.0,        # F  — Código cuenta contable
    14.109375,   # G  — Identificación tercero
    11.44140625, # H  — Sucursal
    11.44140625, # I  — Código producto
    11.88671875, # J  — Código de bodega
    9.6640625,   # K  — Acción
    9.0,         # L  — Cantidad producto
    11.88671875, # M  — Prefijo
    13.33203125, # N  — Consecutivo
    12.44140625, # O  — No. cuota
    12.109375,   # P  — Fecha vencimiento
    13.88671875, # Q  — Código impuesto
    11.44140625, # R  — Código grupo activo fijo
    25.5546875,  # S  — Código activo fijo
    78.88671875, # T  — Descripción
    23.33203125, # U  — Código centro/subcentro de costos
    17.88671875, # V  — Débito
    20.44140625, # W  — Crédito
    17.109375,   # X  — Observaciones
    17.88671875, # Y  — Base gravable libro compras/ventas
    17.109375,   # Z  — Base exenta libro compras/ventas
    # AA (col 27 — Mes de cierre) usa el ancho por defecto
]


def _aplicar_header(ws) -> None:
    """Aplica los estilos de encabezado del template SIIGO a la fila 1."""
    alin = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, len(ENCABEZADOS_SIIGO) + 1):
        color = _COLOR_ROJO_REQUERIDO if col in _COLS_REQUERIDAS else _COLOR_AZUL_OPCIONAL
        fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c = ws.cell(row=1, column=col)
        c.fill      = fill
        c.font      = Font(bold=False)
        c.alignment = alin
    ws.row_dimensions[1].height = 30.0


def _aplicar_anchos(ws) -> None:
    """Aplica los anchos de columna del template SIIGO."""
    for i, ancho in enumerate(_ANCHOS_TEMPLATE, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _escribir_chunk(filas: list[FilaSiigo], filepath: Path) -> None:
    """Escribe un chunk de filas en un archivo Excel con el formato del template SIIGO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Fila 1 — encabezados
    ws.append(ENCABEZADOS_SIIGO)
    _aplicar_header(ws)
    _aplicar_anchos(ws)

    fill_pendiente   = PatternFill(start_color=_COLOR_PENDIENTE,   end_color=_COLOR_PENDIENTE,   fill_type="solid")
    fill_advertencia = PatternFill(start_color=_COLOR_ADVERTENCIA, end_color=_COLOR_ADVERTENCIA, fill_type="solid")

    # Índices de columnas Débito y Crédito (base-1, posiciones 22 y 23)
    idx_debito  = ENCABEZADOS_SIIGO.index("D\xe9bito")  + 1
    idx_credito = ENCABEZADOS_SIIGO.index("Cr\xe9dito") + 1

    for fila_num, fila in enumerate(filas, start=2):
        ws.append(fila.a_lista())

        # Formato numérico colombiano para débito y crédito
        for col_idx in (idx_debito, idx_credito):
            celda = ws.cell(row=fila_num, column=col_idx)
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal="right")

        # Colorear según estado
        if fila.es_pendiente:
            for col in range(1, len(ENCABEZADOS_SIIGO) + 1):
                ws.cell(row=fila_num, column=col).fill = fill_pendiente
        elif not fila.nit_tercero:
            for col in range(1, len(ENCABEZADOS_SIIGO) + 1):
                ws.cell(row=fila_num, column=col).fill = fill_advertencia

    ws.freeze_panes = "A2"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(filepath))
    logger.info("Archivo SIIGO generado: %s (%d filas)", filepath, len(filas))


def exportar_siigo(
    preasientos: list[PreasientoContable],
    output_path: str | None = None,
    incluir_pendientes: bool = True,
    max_filas: int = SIIGO_MAX_FILAS_POR_ARCHIVO,
    df_cuentas=None,
) -> list[str]:
    """
    Genera los archivos Excel para importar en SIIGO.

    Args:
        preasientos:        Lista de preasientos a exportar.
        output_path:        Directorio o ruta base de salida.
                            Por defecto usa OUTPUT_DIR de config.
        incluir_pendientes: Si False, omite líneas con cuenta [PENDIENTE].
        max_filas:          Máximo de filas por archivo (default: 500).
        df_cuentas:         DataFrame del maestro de cuentas contables.
                            Se usa para determinar qué cuentas rellenan
                            las columnas de vencimiento (cols 13-16).

    Returns:
        Lista de rutas absolutas de los archivos generados.
    """
    base_dir = Path(output_path or OUTPUT_DIR)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_nombre = f"siigo_comprobantes_{ts}"

    filas = mapear_lote(preasientos, incluir_pendientes=incluir_pendientes, df_cuentas=df_cuentas)

    if not filas:
        raise ValueError("No hay filas para exportar a SIIGO.")

    chunks = partir_en_chunks(filas, max_filas)
    rutas: list[str] = []

    for i, chunk in enumerate(chunks, start=1):
        sufijo = f"_parte{i}" if len(chunks) > 1 else ""
        filepath = base_dir / f"{base_nombre}{sufijo}.xlsx"
        _escribir_chunk(chunk, filepath)
        rutas.append(str(filepath.resolve()))

    return rutas
