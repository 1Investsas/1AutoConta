"""
Importador de la plantilla Excel de Caja General.

Lee una plantilla diligenciada (la que genera ``plantilla_caja``), valida su
estructura y cada movimiento, y RECALCULA el saldo en lugar de confiar en el
valor digitado en Excel (sección 16 de la especificación).

El resultado es un ``ResultadoImportacion`` con:
  - la metadata del encabezado (empresa, cuenta, mes, año, saldo inicial…),
  - los movimientos ya parseados y con saldo recalculado,
  - los errores por fila y los errores generales,

para que la capa web muestre las filas marcadas y permita corregir antes de
guardar definitivamente.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

from app.caja import modelo_caja as mc
from app.caja.plantilla_caja import (
    COL_CENTRO, COL_COMPROBANTE, COL_CONCEPTO, COL_CONTRAPARTIDA, COL_ENTRADA,
    COL_FECHA, COL_NIT, COL_NOMBRE, COL_OBSERVACIONES, COL_SALIDA, COLUMNAS_TABLA,
    FILA_ANIO, FILA_CUENTA, FILA_EMPRESA, FILA_MES, FILA_RESPONSABLE,
    FILA_SALDO_INICIAL, FILA_TABLA_HEADER, FILA_TABLA_INICIO, FILA_VERSION,
    HOJA_CAJA, VERSION_PLANTILLA,
)

logger = logging.getLogger(__name__)


@dataclass
class ResultadoImportacion:
    """Resultado del parseo + validación de una plantilla de caja."""

    empresa: str = ""
    cuenta_caja: str = ""
    anio: Optional[int] = None
    mes: Optional[int] = None
    saldo_inicial: Decimal = field(default_factory=lambda: Decimal("0"))
    responsable: str = ""
    version: str = ""
    movimientos: list[mc.MovimientoCaja] = field(default_factory=list)
    errores_por_fila: dict[int, list[str]] = field(default_factory=dict)
    errores_generales: list[str] = field(default_factory=list)

    @property
    def tiene_errores(self) -> bool:
        return bool(self.errores_generales) or any(self.errores_por_fila.values())

    @property
    def n_errores(self) -> int:
        return len(self.errores_generales) + sum(
            1 for errs in self.errores_por_fila.values() if errs
        )


def _celda(ws, fila: int, col: int):
    return ws.cell(row=fila, column=col).value


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _contrapartida_limpia(valor) -> str:
    """Ignora el marcador '(dividida)' que la plantilla escribe para subdivisiones."""
    txt = _texto(valor)
    return "" if txt.lower() == "(dividida)" else txt


def _parse_mes(valor) -> Optional[int]:
    """Lee el mes desde el encabezado (acepta '6', '06 — Junio', 'Junio')."""
    txt = _texto(valor)
    if not txt:
        return None
    cabeza = txt.split("—")[0].split("-")[0].strip()
    if cabeza.isdigit():
        m = int(cabeza)
        return m if 1 <= m <= 12 else None
    for i, nombre in enumerate(mc.MESES_ES):
        if nombre and nombre.lower() in txt.lower():
            return i
    return None


def importar_plantilla(path: str | Path) -> ResultadoImportacion:
    """Lee y valida una plantilla de caja diligenciada.

    Args:
        path: Ruta al archivo .xlsx diligenciado.

    Returns:
        ResultadoImportacion con metadata, movimientos (saldo recalculado) y
        errores. No lanza excepción por errores de diligenciamiento; sí puede
        lanzar si el archivo no es un .xlsx legible.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[HOJA_CAJA] if HOJA_CAJA in wb.sheetnames else wb.active

    res = ResultadoImportacion()

    # ── Validación de estructura ────────────────────────────────────────────
    encabezados = [_texto(_celda(ws, FILA_TABLA_HEADER, c + 1))
                   for c in range(len(COLUMNAS_TABLA))]
    if encabezados[:3] != COLUMNAS_TABLA[:3]:
        res.errores_generales.append(
            "El archivo no corresponde a la plantilla de Caja General vigente "
            "(no se reconoce la tabla de movimientos)."
        )
        return res

    # ── Metadata del encabezado ─────────────────────────────────────────────
    res.empresa = _texto(_celda(ws, FILA_EMPRESA, 2))
    res.cuenta_caja = _texto(_celda(ws, FILA_CUENTA, 2))
    res.mes = _parse_mes(_celda(ws, FILA_MES, 2))
    anio_raw = _texto(_celda(ws, FILA_ANIO, 2))
    res.anio = int(anio_raw) if anio_raw.isdigit() else None
    res.saldo_inicial = mc.a_decimal(_celda(ws, FILA_SALDO_INICIAL, 2))
    res.responsable = _texto(_celda(ws, FILA_RESPONSABLE, 2))
    res.version = _texto(_celda(ws, FILA_VERSION, 2))

    if res.version and res.version != VERSION_PLANTILLA:
        res.errores_generales.append(
            f"La plantilla es versión {res.version}; la versión vigente es "
            f"{VERSION_PLANTILLA}. Descarga una plantilla nueva si hay problemas."
        )

    # ── Tabla de movimientos ────────────────────────────────────────────────
    movimientos: list[mc.MovimientoCaja] = []
    filas_vacias_seguidas = 0
    fila = FILA_TABLA_INICIO
    seq = 0
    while filas_vacias_seguidas < 15:
        valores = {
            "movement_date": _celda(ws, fila, COL_FECHA),
            "comprobante": _celda(ws, fila, COL_COMPROBANTE),
            "concept": _celda(ws, fila, COL_CONCEPTO),
            "third_party_nit": _celda(ws, fila, COL_NIT),
            "third_party_name": _celda(ws, fila, COL_NOMBRE),
            "cost_center": _celda(ws, fila, COL_CENTRO),
            "contrapartida": _contrapartida_limpia(_celda(ws, fila, COL_CONTRAPARTIDA)),
            "inflow_amount": _celda(ws, fila, COL_ENTRADA),
            "outflow_amount": _celda(ws, fila, COL_SALIDA),
            "observations": _celda(ws, fila, COL_OBSERVACIONES),
        }
        if _fila_vacia(valores):
            filas_vacias_seguidas += 1
            fila += 1
            continue
        filas_vacias_seguidas = 0
        seq += 1
        valores["sequence"] = seq
        mov = mc.desde_dict(valores)
        movimientos.append(mov)
        errores = mc.validar_movimiento(mov, res.anio, res.mes)
        if errores:
            res.errores_por_fila[fila] = errores
        fila += 1

    # Recalcular el saldo: no se confía en el valor digitado en Excel.
    res.movimientos = mc.recalcular_saldos(movimientos, res.saldo_inicial)
    mc.renumerar(res.movimientos)
    return res


def _fila_vacia(valores: dict) -> bool:
    """True si la fila no tiene ningún dato diligenciable."""
    relevantes = (
        "movement_date", "comprobante", "concept", "third_party_nit",
        "third_party_name", "cost_center", "contrapartida", "inflow_amount",
        "outflow_amount", "observations",
    )
    return all(_texto(valores.get(k)) == "" for k in relevantes)
