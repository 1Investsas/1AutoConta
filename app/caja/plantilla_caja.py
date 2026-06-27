"""
Plantilla Excel del módulo Caja General.

Genera el archivo de trabajo en el que el usuario diligencia (o consulta) los
movimientos de efectivo de un período mensual de caja. A diferencia de Bancos,
aquí la aplicación *estructura* el formato; no se importa un extracto externo.

Dos modos:
  - Plantilla **vacía** (``generar_plantilla``): lista para diligenciar a mano.
  - Plantilla **prediligenciada** (``generar_plantilla`` con ``movimientos``):
    incluye los movimientos ya registrados en la app, como respaldo / soporte.

Características (secciones 17 y 18 de la especificación):
  - Encabezado con la información general del período.
  - Tabla de movimientos con validación de tipo (Entrada/Salida) y de fecha.
  - Columna **Saldo** calculada por fórmula y protegida (no editable).
  - Formato monetario en Entrada, Salida y Saldo.
  - Hoja auxiliar ``Terceros`` para autocompletar NIT ↔ nombre en Excel.
  - Versión de plantilla para control de compatibilidad.

El layout es de posiciones fijas (ver constantes) para que el importador lo
lea de forma determinista.
"""

from __future__ import annotations

import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.caja.modelo_caja import ENTRADA, SALIDA, MESES_ES

# Versión del formato. Súbela si cambia la estructura (el importador la valida).
VERSION_PLANTILLA = "1.0"

HOJA_CAJA = "Caja General"
HOJA_TERCEROS = "Terceros"

# ── Layout del encabezado (fila → etiqueta). Los valores van en la columna B ──
FILA_TITULO = 1
FILA_EMPRESA = 3
FILA_CUENTA = 4
FILA_MES = 5
FILA_ANIO = 6
FILA_SALDO_INICIAL = 7
FILA_RESPONSABLE = 8
FILA_FECHA_GEN = 9
FILA_VERSION = 10

ETIQUETAS_ENCABEZADO = {
    FILA_EMPRESA: "Empresa",
    FILA_CUENTA: "Cuenta de caja",
    FILA_MES: "Mes",
    FILA_ANIO: "Año",
    FILA_SALDO_INICIAL: "Saldo inicial",
    FILA_RESPONSABLE: "Responsable",
    FILA_FECHA_GEN: "Fecha de generación",
    FILA_VERSION: "Versión plantilla",
}

# ── Tabla de movimientos ─────────────────────────────────────────────────────
FILA_TABLA_HEADER = 12
FILA_TABLA_INICIO = 13
FILAS_VACIAS_PLANTILLA = 60  # filas en blanco listas para diligenciar

COLUMNAS_TABLA = [
    "Fecha", "Tipo movimiento", "Concepto", "NIT tercero", "Nombre tercero",
    "Centro de costo", "Categoría / cuenta", "Entrada", "Salida", "Saldo",
    "Observaciones",
]
# Índices (1-based) de columnas clave dentro de la tabla.
COL_FECHA = 1
COL_TIPO = 2
COL_CONCEPTO = 3
COL_NIT = 4
COL_NOMBRE = 5
COL_CENTRO = 6
COL_CATEGORIA = 7
COL_ENTRADA = 8
COL_SALIDA = 9
COL_SALDO = 10
COL_OBSERVACIONES = 11

_FMT_MONEDA = '#,##0'
_FMT_FECHA = 'DD/MM/YYYY'

_AZUL = "1F4E78"
_AZUL_CLARO = "DDEBF7"
_GRIS = "F2F2F2"
_AMARILLO = "FFF2CC"


def _borde_fino() -> Border:
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def generar_plantilla(
    *,
    empresa: str = "",
    cuenta_caja: str = "",
    anio: Optional[int] = None,
    mes: Optional[int] = None,
    saldo_inicial="0",
    responsable: str = "",
    movimientos: Optional[list[dict]] = None,
    terceros: Optional[list[dict]] = None,
) -> bytes:
    """Construye la plantilla Excel y la retorna como bytes (.xlsx).

    Args:
        empresa:        Nombre de la empresa propietaria.
        cuenta_caja:    Nombre de la cuenta de caja.
        anio, mes:      Período del registro.
        saldo_inicial:  Saldo inicial del mes.
        responsable:    Persona a cargo del registro.
        movimientos:    Si se pasan, genera la plantilla PREDILIGENCIADA con
                        estos movimientos (lista de dicts a_dict()). Si es None
                        o vacía, genera la plantilla VACÍA para diligenciar.
        terceros:       Lista de {'nit','nombre'} para la hoja auxiliar Terceros.

    Returns:
        Contenido del archivo .xlsx en bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = HOJA_CAJA

    _escribir_encabezado(ws, empresa, cuenta_caja, anio, mes, saldo_inicial, responsable)
    _escribir_tabla_header(ws)
    n_filas = _escribir_movimientos(ws, movimientos or [])
    _aplicar_validaciones(ws, anio, mes)
    _aplicar_anchos(ws)
    _proteger(ws, n_filas)

    _escribir_hoja_terceros(wb, terceros or [])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _escribir_encabezado(ws, empresa, cuenta_caja, anio, mes, saldo_inicial, responsable):
    """Escribe el bloque de información general del período."""
    titulo = ws.cell(row=FILA_TITULO, column=1, value="CAJA GENERAL — MOVIMIENTOS DE EFECTIVO")
    titulo.font = Font(bold=True, size=14, color=_AZUL)

    mes_txt = f"{mes} — {MESES_ES[mes]}" if mes and 1 <= mes <= 12 else ""
    valores = {
        FILA_EMPRESA: empresa,
        FILA_CUENTA: cuenta_caja,
        FILA_MES: mes_txt,
        FILA_ANIO: anio or "",
        FILA_SALDO_INICIAL: float(saldo_inicial or 0),
        FILA_RESPONSABLE: responsable,
        FILA_FECHA_GEN: date.today().strftime("%d/%m/%Y"),
        FILA_VERSION: VERSION_PLANTILLA,
    }
    relleno = PatternFill(start_color=_GRIS, end_color=_GRIS, fill_type="solid")
    relleno_val = PatternFill(start_color=_AMARILLO, end_color=_AMARILLO, fill_type="solid")
    for fila, etiqueta in ETIQUETAS_ENCABEZADO.items():
        c_lab = ws.cell(row=fila, column=1, value=etiqueta)
        c_lab.font = Font(bold=True)
        c_lab.fill = relleno
        c_val = ws.cell(row=fila, column=2, value=valores.get(fila, ""))
        c_val.fill = relleno_val
        if fila == FILA_SALDO_INICIAL:
            c_val.number_format = _FMT_MONEDA


def _escribir_tabla_header(ws):
    """Escribe la fila de encabezados de la tabla de movimientos."""
    fill = PatternFill(start_color=_AZUL, end_color=_AZUL, fill_type="solid")
    fuente = Font(bold=True, color="FFFFFF")
    alin = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, titulo in enumerate(COLUMNAS_TABLA, start=1):
        c = ws.cell(row=FILA_TABLA_HEADER, column=col, value=titulo)
        c.fill = fill
        c.font = fuente
        c.alignment = alin
        c.border = _borde_fino()
    ws.row_dimensions[FILA_TABLA_HEADER].height = 28
    ws.freeze_panes = f"A{FILA_TABLA_INICIO}"


def _formula_saldo(fila: int) -> str:
    """Fórmula del saldo acumulado para una fila de datos (protegida).

    Saldo = saldo inicial + Σ entradas − Σ salidas hasta esta fila. Las filas
    sin entrada ni salida quedan en blanco para no arrastrar el saldo inicial.
    """
    saldo_ini = f"$B${FILA_SALDO_INICIAL}"
    ent = get_column_letter(COL_ENTRADA)
    sal = get_column_letter(COL_SALIDA)
    rango_ent = f"${ent}${FILA_TABLA_INICIO}:${ent}{fila}"
    rango_sal = f"${sal}${FILA_TABLA_INICIO}:${sal}{fila}"
    return (
        f'=IF(AND({ent}{fila}="",{sal}{fila}=""),"",'
        f'{saldo_ini}+SUM({rango_ent})-SUM({rango_sal}))'
    )


def _escribir_movimientos(ws, movimientos: list[dict]) -> int:
    """Escribe las filas de movimientos (o filas vacías) y retorna cuántas filas hay."""
    n = max(len(movimientos), FILAS_VACIAS_PLANTILLA)
    tipo_label = {ENTRADA: "Entrada", SALIDA: "Salida"}
    for i in range(n):
        fila = FILA_TABLA_INICIO + i
        mov = movimientos[i] if i < len(movimientos) else None

        if mov:
            fecha = mov.get("movement_date") or ""
            ws.cell(row=fila, column=COL_FECHA, value=fecha)
            ws.cell(row=fila, column=COL_TIPO,
                    value=tipo_label.get(mov.get("movement_type", ""), ""))
            ws.cell(row=fila, column=COL_CONCEPTO, value=mov.get("concept", ""))
            ws.cell(row=fila, column=COL_NIT, value=mov.get("third_party_nit", ""))
            ws.cell(row=fila, column=COL_NOMBRE, value=mov.get("third_party_name", ""))
            ws.cell(row=fila, column=COL_CENTRO, value=mov.get("cost_center", ""))
            ws.cell(row=fila, column=COL_CATEGORIA, value=mov.get("category", ""))
            ent = float(mov.get("inflow_amount") or 0)
            sal = float(mov.get("outflow_amount") or 0)
            ws.cell(row=fila, column=COL_ENTRADA, value=ent or None)
            ws.cell(row=fila, column=COL_SALIDA, value=sal or None)
            ws.cell(row=fila, column=COL_OBSERVACIONES, value=mov.get("observations", ""))

        # Saldo siempre por fórmula (también en filas vacías → quedan en blanco).
        c_saldo = ws.cell(row=fila, column=COL_SALDO, value=_formula_saldo(fila))
        c_saldo.number_format = _FMT_MONEDA

        # Formato de las columnas numéricas y de fecha + bordes.
        ws.cell(row=fila, column=COL_FECHA).number_format = _FMT_FECHA
        ws.cell(row=fila, column=COL_ENTRADA).number_format = _FMT_MONEDA
        ws.cell(row=fila, column=COL_SALIDA).number_format = _FMT_MONEDA
        for col in range(1, len(COLUMNAS_TABLA) + 1):
            ws.cell(row=fila, column=col).border = _borde_fino()
    return n


def _aplicar_validaciones(ws, anio, mes):
    """Agrega listas desplegables y validación de fecha a la tabla."""
    fila_fin = FILA_TABLA_INICIO + max(FILAS_VACIAS_PLANTILLA, 1) - 1

    # Tipo de movimiento: lista controlada.
    dv_tipo = DataValidation(
        type="list", formula1='"Entrada,Salida"', allow_blank=True,
        showErrorMessage=True,
    )
    dv_tipo.error = "El tipo debe ser Entrada o Salida."
    dv_tipo.errorTitle = "Tipo de movimiento inválido"
    col_tipo = get_column_letter(COL_TIPO)
    dv_tipo.add(f"{col_tipo}{FILA_TABLA_INICIO}:{col_tipo}{fila_fin}")
    ws.add_data_validation(dv_tipo)

    # Fecha dentro del período, si se conoce el mes/año.
    if anio and mes:
        ini = date(anio, mes, 1)
        fin = date(anio + (mes == 12), (mes % 12) + 1, 1)
        dv_fecha = DataValidation(
            type="date", operator="between",
            formula1=ini.strftime("%Y-%m-%d"), formula2=fin.strftime("%Y-%m-%d"),
            allow_blank=True, showErrorMessage=True,
        )
        dv_fecha.error = f"La fecha debe pertenecer a {mes:02d}/{anio}."
        dv_fecha.errorTitle = "Fecha fuera del período"
        col_f = get_column_letter(COL_FECHA)
        dv_fecha.add(f"{col_f}{FILA_TABLA_INICIO}:{col_f}{fila_fin}")
        ws.add_data_validation(dv_fecha)


def _aplicar_anchos(ws):
    anchos = [13, 16, 34, 14, 30, 18, 22, 15, 15, 16, 30]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _proteger(ws, n_filas: int):
    """Protege la hoja dejando editables solo las celdas diligenciables.

    Se bloquea la columna Saldo (fórmula) y el encabezado; el resto de la tabla
    y el valor del saldo inicial quedan desbloqueados para que el usuario los
    pueda diligenciar.
    """
    # Desbloquear el valor del saldo inicial.
    ws.cell(row=FILA_SALDO_INICIAL, column=2).protection = Protection(locked=False)

    fila_fin = FILA_TABLA_INICIO + max(n_filas, FILAS_VACIAS_PLANTILLA) - 1
    for fila in range(FILA_TABLA_INICIO, fila_fin + 1):
        for col in range(1, len(COLUMNAS_TABLA) + 1):
            if col == COL_SALDO:
                continue  # saldo: queda bloqueado (fórmula protegida)
            ws.cell(row=fila, column=col).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.enable()
    # Permitir seleccionar/usar autofiltros aun con la hoja protegida.
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    ws.protection.formatCells = False


def _escribir_hoja_terceros(wb, terceros: list[dict]):
    """Crea la hoja auxiliar Terceros (NIT, Nombre) para autocompletar en Excel."""
    ws = wb.create_sheet(HOJA_TERCEROS)
    fill = PatternFill(start_color=_AZUL_CLARO, end_color=_AZUL_CLARO, fill_type="solid")
    for col, titulo in enumerate(["NIT", "Nombre tercero"], start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = Font(bold=True)
        c.fill = fill
    for i, t in enumerate(terceros, start=2):
        ws.cell(row=i, column=1, value=str(t.get("nit", "")))
        ws.cell(row=i, column=2, value=str(t.get("nombre", "")))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    # Nota de uso para el usuario.
    fila_nota = max(len(terceros) + 3, 4)
    ws.cell(
        row=fila_nota, column=1,
        value="Hoja de apoyo: úsala con BUSCARV/XLOOKUP para autocompletar NIT ↔ nombre.",
    ).font = Font(italic=True, color="808080")
