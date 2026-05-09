"""
Exportador de resultados a Excel formateado.

Genera un archivo .xlsx con cuatro pestañas:
1. Resumen    — estadísticas del proceso.
2. Preasientos — todas las líneas contables listas para importar.
3. Excepciones — documentos con errores o sin clasificar.
4. Bitácora   — registro cronológico de acciones del proceso.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import PreasientoContable, RegistroBitacora

logger = logging.getLogger(__name__)

# Colores corporativos (azul oscuro)
COLOR_HEADER = "1F3864"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_PENDIENTE = "FF0000"   # Rojo para cuentas pendientes
COLOR_ADVERTENCIA = "FFC000"  # Amarillo para advertencias
COLOR_OK = "E2EFDA"           # Verde claro para filas sin problemas


def _estilo_header(ws, fila: int, num_cols: int) -> None:
    """Aplica estilo de encabezado a una fila completa."""
    fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    alin = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lado = Side(style="thin", color="000000")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)

    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = fill
        celda.font = font
        celda.alignment = alin
        celda.border = borde


def _autoajustar_columnas(ws) -> None:
    """Ajusta el ancho de cada columna según el contenido máximo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                largo = len(str(cell.value)) if cell.value is not None else 0
                if largo > max_len:
                    max_len = largo
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _formato_moneda(valor: float) -> str:
    """Formatea un valor numérico como moneda colombiana."""
    return f"${valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def exportar_excel(
    preasientos: list[PreasientoContable],
    excepciones: list[dict],
    bitacora: list[RegistroBitacora],
    output_path: str,
    archivo_origen: Optional[str] = None,
) -> str:
    """
    Genera el archivo Excel con los resultados del proceso.

    Args:
        preasientos:   Lista de PreasientoContable generados.
        excepciones:   Lista de dicts describiendo documentos con problemas.
        bitacora:      Lista de RegistroBitacora de la sesión.
        output_path:   Ruta del directorio o archivo de salida.
        archivo_origen: Nombre del archivo RADIAN procesado (para el resumen).

    Returns:
        Ruta absoluta del archivo Excel generado.
    """
    # Determinar ruta de salida
    path = Path(output_path)
    if path.is_dir() or not path.suffix:
        path.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = path / f"preasientos_{ts}.xlsx"
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        filepath = path

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Pestaña 1: Resumen
    # ------------------------------------------------------------------
    ws_res = wb.active
    ws_res.title = "Resumen"
    _generar_resumen(ws_res, preasientos, excepciones, archivo_origen)

    # ------------------------------------------------------------------
    # Pestaña 2: Preasientos
    # ------------------------------------------------------------------
    ws_pre = wb.create_sheet("Preasientos")
    _generar_preasientos(ws_pre, preasientos)

    # ------------------------------------------------------------------
    # Pestaña 3: Excepciones
    # ------------------------------------------------------------------
    ws_exc = wb.create_sheet("Excepciones")
    _generar_excepciones(ws_exc, excepciones)

    # ------------------------------------------------------------------
    # Pestaña 4: Bitácora
    # ------------------------------------------------------------------
    ws_bit = wb.create_sheet("Bitácora")
    _generar_bitacora(ws_bit, bitacora)

    wb.save(str(filepath))
    logger.info("Excel exportado: %s", filepath)

    # En modo cloud, subir a Azure Blob Storage
    from app.storage import is_cloud, save_local_file
    if is_cloud():
        ref = save_local_file(str(filepath), "output", filepath.name)
        return ref

    return str(filepath)


def _generar_resumen(
    ws,
    preasientos: list[PreasientoContable],
    excepciones: list[dict],
    archivo_origen: Optional[str],
) -> None:
    """Rellena la pestaña Resumen."""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    font_titulo = Font(bold=True, size=14, color=COLOR_HEADER)
    font_clave = Font(bold=True, size=11)

    ws["A1"] = "RESUMEN DE PROCESAMIENTO CONTABLE"
    ws["A1"].font = font_titulo
    ws.merge_cells("A1:B1")

    datos = [
        ("Fecha de procesamiento", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Archivo origen", archivo_origen or "N/D"),
        ("Total documentos", len(preasientos)),
        ("Documentos con excepciones", len(excepciones)),
        ("Documentos OK", len(preasientos) - len(excepciones)),
        ("", ""),
    ]

    # Conteo por clasificación
    conteo: dict[str, int] = {}
    for p in preasientos:
        conteo[p.clasificacion] = conteo.get(p.clasificacion, 0) + 1

    datos.append(("DESGLOSE POR TIPO", ""))
    for clase, cant in sorted(conteo.items()):
        datos.append((clase, cant))

    for i, (clave, valor) in enumerate(datos, start=2):
        celda_c = ws.cell(row=i, column=1, value=clave)
        celda_v = ws.cell(row=i, column=2, value=valor)
        if clave and clave == clave.upper() and clave != "":
            celda_c.font = font_clave
        celda_c.alignment = Alignment(horizontal="left")
        celda_v.alignment = Alignment(horizontal="right")


def _generar_preasientos(ws, preasientos: list[PreasientoContable]) -> None:
    """Rellena la pestaña Preasientos con todas las líneas contables."""
    encabezados = [
        "CUFE/CUDE", "Tipo Documento", "Clasificación",
        "Cód. Comprobante", "Título Comprobante",
        "Fecha Emisión", "Folio", "Prefijo",
        "Tercero NIT", "Tercero Nombre", "Tercero Encontrado",
        "# Línea", "Cuenta", "Descripción Cuenta",
        "Débito", "Crédito", "Concepto", "Cuenta Pendiente",
    ]

    ws.append(encabezados)
    _estilo_header(ws, 1, len(encabezados))

    fill_pendiente = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    fill_advertencia = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_rojo = Font(color=COLOR_PENDIENTE, bold=True)

    fila = 2
    for p in preasientos:
        fecha_str = p.fecha_emision.strftime("%d/%m/%Y") if p.fecha_emision else ""
        for linea in p.lineas:
            ws.append([
                p.cufe,
                p.tipo_documento,
                p.clasificacion,
                p.codigo_comprobante,
                p.titulo_comprobante,
                fecha_str,
                p.folio,
                p.prefijo,
                p.tercero_nit,
                p.tercero_nombre,
                "Sí" if p.tercero_encontrado else "No",
                linea.numero_linea,
                linea.cuenta,
                linea.descripcion_cuenta,
                linea.debito,
                linea.credito,
                linea.concepto,
                "Sí" if linea.es_pendiente else "",
            ])

            # Colorear filas con cuenta pendiente
            if linea.es_pendiente:
                for col in range(1, len(encabezados) + 1):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = fill_pendiente
                ws.cell(row=fila, column=13).font = font_rojo
            elif not p.tercero_encontrado:
                for col in range(1, len(encabezados) + 1):
                    ws.cell(row=fila, column=col).fill = fill_advertencia

            # Formato numérico colombiano para débitos y créditos
            for col_num in (15, 16):
                celda = ws.cell(row=fila, column=col_num)
                celda.number_format = '#,##0.00'
                celda.alignment = Alignment(horizontal="right")

            fila += 1

    ws.freeze_panes = "A2"
    _autoajustar_columnas(ws)


def _generar_excepciones(ws, excepciones: list[dict]) -> None:
    """Rellena la pestaña Excepciones."""
    encabezados = ["CUFE/CUDE", "Tipo Documento", "Clasificación",
                   "NIT Tercero", "Total", "Errores"]
    ws.append(encabezados)
    _estilo_header(ws, 1, len(encabezados))

    for exc in excepciones:
        ws.append([
            exc.get("cufe", ""),
            exc.get("tipo_documento", ""),
            exc.get("clasificacion", ""),
            exc.get("tercero_nit", ""),
            exc.get("total", 0.0),
            "; ".join(exc.get("errores", [])),
        ])

    _autoajustar_columnas(ws)


def _generar_bitacora(ws, bitacora: list[RegistroBitacora]) -> None:
    """Rellena la pestaña Bitácora."""
    encabezados = ["Timestamp", "Nivel", "Módulo", "Acción", "Detalle", "CUFE"]
    ws.append(encabezados)
    _estilo_header(ws, 1, len(encabezados))

    fill_error = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    fill_warn = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for i, reg in enumerate(bitacora, start=2):
        ws.append([
            reg.timestamp.strftime("%d/%m/%Y %H:%M:%S"),
            reg.nivel,
            reg.modulo,
            reg.accion,
            reg.detalle,
            reg.cufe or "",
        ])
        if reg.nivel == "ERROR":
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = fill_error
        elif reg.nivel == "WARNING":
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = fill_warn

    _autoajustar_columnas(ws)
