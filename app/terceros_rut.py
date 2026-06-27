"""
Actualización del maestro de terceros a partir de RUTs de la DIAN.

Toma los datos leídos de uno o varios RUT (ver ``app.rut``) y los inserta o
actualiza en el archivo ``Listado_de_Terceros.xlsx`` —el mismo maestro que usa
el pipeline RADIAN para cruzar terceros—. El upsert:

- Conserva el formato del archivo existente (filas 1–6 informativas y los
  encabezados en la fila 7) trabajando con ``openpyxl`` sobre el libro real.
- Reconoce las columnas por su nombre (sin distinguir mayúsculas ni tildes),
  de modo que funciona aunque el export del sistema use variantes.
- Hace *match* por identificación (NIT/cédula, solo dígitos): si el tercero ya
  existe lo actualiza; si no, lo agrega.
- Si el archivo no existe todavía, crea uno nuevo con un conjunto de columnas
  por defecto y los encabezados en la fila 7 (compatible con el lector).
"""

from __future__ import annotations

import io
import logging
import unicodedata
from typing import Optional

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

# Fila (1-based de Excel) donde van los encabezados de columnas. Coincide con
# ``config.FILA_ENCABEZADOS_MAESTROS`` (header=6 en pandas → fila 7 de Excel).
FILA_ENCABEZADOS_EXCEL = 7


# ---------------------------------------------------------------------------
# Definición de campos: campo canónico → (encabezado por defecto, alias)
# ---------------------------------------------------------------------------
# Los ``alias`` se comparan (normalizados) contra los encabezados del archivo
# existente para ubicar cada columna. El ``encabezado`` se usa al crear un
# archivo nuevo.
_CAMPOS: list[tuple[str, str, tuple[str, ...]]] = [
    ("nombre",              "Nombre tercero",
     ("nombre tercero", "nombre", "razon social", "nombre o razon social",
      "nombre completo", "tercero", "nombre del tercero")),
    ("tipo_identificacion", "Tipo de identificación",
     ("tipo de identificacion", "tipo identificacion", "tipo de documento",
      "tipo documento", "tipo doc", "tipo id")),
    ("identificacion",      "Identificación",
     ("identificacion", "nit", "numero de identificacion", "numero identificacion",
      "nro identificacion", "documento", "cedula", "no identificacion",
      "numero de documento")),
    ("dv",                  "Digito verificación",
     ("digito verificacion", "dv", "digito de verificacion", "digito",
      "digito verificación")),
    ("regimen_iva",         "Tipo de regimen IVA",
     ("tipo de regimen iva", "regimen iva", "regimen", "responsabilidad iva",
      "responsable de iva", "tipo de régimen iva")),
    ("direccion",           "Dirección",
     ("direccion", "direccion principal", "dir")),
    ("ciudad",              "Ciudad",
     ("ciudad", "ciudad/municipio", "ciudad municipio", "municipio")),
    ("departamento",        "Departamento",
     ("departamento", "depto")),
    ("pais",                "País",
     ("pais",)),
    ("telefono",            "Teléfono",
     ("telefono", "telefono 1", "tel", "celular", "movil", "teléfono")),
    ("correo",              "Correo electrónico",
     ("correo electronico", "correo", "email", "e-mail", "mail",
      "correo electrónico")),
    ("primer_apellido",     "Primer apellido",  ("primer apellido",)),
    ("segundo_apellido",    "Segundo apellido", ("segundo apellido",)),
    ("primer_nombre",       "Primer nombre",    ("primer nombre",)),
    ("segundo_nombre",      "Segundo nombre",   ("segundo nombre", "otros nombres")),
    ("estado",              "Estado",           ("estado",)),
]

# Columnas por defecto al crear un maestro nuevo (orden de presentación).
_COLUMNAS_NUEVO = [
    "Nombre tercero", "Tipo de identificación", "Identificación",
    "Digito verificación", "Sucursal", "Tipo de regimen IVA",
    "Dirección", "Ciudad", "Departamento", "País",
    "Teléfono", "Correo electrónico", "Nombres contacto", "Estado",
]

# Mapa rápido alias→campo y campo→encabezado por defecto.
_ALIAS_A_CAMPO: dict[str, str] = {}
for _campo, _enc, _aliases in _CAMPOS:
    for _a in _aliases:
        _ALIAS_A_CAMPO[_a] = _campo
_CAMPO_A_ENCABEZADO = {c: e for c, e, _ in _CAMPOS}


def _normalizar(texto: object) -> str:
    """Normaliza un texto para comparar encabezados: minúsculas, sin tildes."""
    s = str(texto or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(s.split())


def _solo_digitos(valor: object) -> str:
    """Deja solo los dígitos de un identificador (igual que el importador)."""
    return "".join(c for c in str(valor or "") if c.isdigit())


def mapear_rut_a_tercero(rut: dict) -> dict:
    """Convierte el dict leído de un RUT en un tercero con claves canónicas."""
    return {
        "identificacion":      _solo_digitos(rut.get("nit", "")),
        "dv":                  rut.get("dv", ""),
        "tipo_identificacion": rut.get("tipo_identificacion", ""),
        "nombre":              rut.get("nombre", ""),
        "primer_apellido":     rut.get("primer_apellido", ""),
        "segundo_apellido":    rut.get("segundo_apellido", ""),
        "primer_nombre":       rut.get("primer_nombre", ""),
        "segundo_nombre":      rut.get("otros_nombres", ""),
        "direccion":           rut.get("direccion", ""),
        "ciudad":              rut.get("ciudad", ""),
        "departamento":        rut.get("departamento", ""),
        "pais":                rut.get("pais", ""),
        "telefono":            rut.get("telefono", ""),
        "correo":              rut.get("correo", ""),
        "regimen_iva":         rut.get("regimen_iva", ""),
        "estado":              "Activo",
    }


def _abrir_o_crear(contenido: Optional[bytes]):
    """Devuelve (workbook, worksheet, fila_encabezados, creado_nuevo)."""
    if contenido:
        wb = load_workbook(io.BytesIO(contenido))
        ws = wb.active
        fila_enc = _detectar_fila_encabezados(ws)
        return wb, ws, fila_enc, False

    # Crear un maestro nuevo con encabezados en la fila 7.
    wb = Workbook()
    ws = wb.active
    ws.title = "Terceros"
    ws.cell(row=1, column=1, value="Listado de Terceros")
    for col, nombre in enumerate(_COLUMNAS_NUEVO, start=1):
        ws.cell(row=FILA_ENCABEZADOS_EXCEL, column=col, value=nombre)
    return wb, ws, FILA_ENCABEZADOS_EXCEL, True


def _detectar_fila_encabezados(ws) -> int:
    """Ubica la fila de encabezados buscando la columna de identificación.

    El export del sistema pone los encabezados en la fila 7; si por alguna razón
    están en otra fila (entre las primeras 15), se detecta buscando una celda
    cuyo texto coincida con un alias de «Identificación».
    """
    aliases_id = {a for a, c in _ALIAS_A_CAMPO.items() if c == "identificacion"}
    max_scan = min(15, ws.max_row or 15)
    for fila in range(1, max_scan + 1):
        for celda in ws[fila]:
            if _normalizar(celda.value) in aliases_id:
                return fila
    return FILA_ENCABEZADOS_EXCEL


def _mapa_columnas(ws, fila_enc: int) -> dict[str, int]:
    """Construye {campo_canónico: índice_de_columna} desde la fila de encabezados."""
    mapa: dict[str, int] = {}
    for celda in ws[fila_enc]:
        campo = _ALIAS_A_CAMPO.get(_normalizar(celda.value))
        if campo and campo not in mapa:
            mapa[campo] = celda.column
    return mapa


def actualizar_maestro_terceros(
    terceros: list[dict],
    contenido: Optional[bytes] = None,
) -> tuple[bytes, dict]:
    """Inserta/actualiza terceros en el maestro y devuelve (bytes, resumen).

    Args:
        terceros:  Lista de terceros con claves canónicas
                   (ver ``mapear_rut_a_tercero``).
        contenido: Bytes del ``Listado_de_Terceros.xlsx`` existente, o ``None``
                   para crear uno nuevo.

    Returns:
        Tupla ``(bytes_actualizados, resumen)``. ``resumen`` incluye
        ``agregados``, ``actualizados``, ``detalle`` (por tercero) y
        ``columnas`` (campos detectados en el archivo).
    """
    wb, ws, fila_enc, creado = _abrir_o_crear(contenido)
    columnas = _mapa_columnas(ws, fila_enc)

    if "identificacion" not in columnas:
        # El archivo de terceros no tiene una columna de identificación. La causa
        # más común es haber subido otro maestro (p. ej. el Plan de Cuentas) en
        # la casilla de Terceros: damos un mensaje claro según lo que parezca ser.
        from app.maestros import clasificar_encabezados, ETIQUETA_MAESTRO
        encabezados = [c.value for c in ws[fila_enc]]
        clase = clasificar_encabezados([str(e) for e in encabezados if e])
        if clase in ("cuentas", "comprobantes"):
            etiqueta = ETIQUETA_MAESTRO.get(clase, clase)
            raise ValueError(
                f"El archivo guardado como «Listado de Terceros» parece ser el "
                f"«{etiqueta}». Ve a Configuraciones → Empresas → Maestros y sube "
                f"el archivo de terceros correcto en su casilla."
            )
        raise ValueError(
            "El maestro de terceros no tiene una columna de «Identificación» "
            "(NIT/Cédula) reconocible en la fila 7. Verifica que el archivo sea "
            "el Listado de Terceros exportado del sistema."
        )

    col_id = columnas["identificacion"]

    # Índice de terceros existentes: identificación (solo dígitos) → nº de fila.
    indice: dict[str, int] = {}
    ultima_fila = fila_enc
    for fila in range(fila_enc + 1, (ws.max_row or fila_enc) + 1):
        valor = ws.cell(row=fila, column=col_id).value
        ident = _solo_digitos(valor)
        if ident:
            indice[ident] = fila
            ultima_fila = fila
        elif any(c.value not in (None, "") for c in ws[fila]):
            # Fila con datos pero sin identificación: la contamos para no
            # sobrescribirla al agregar nuevas filas.
            ultima_fila = fila

    agregados = actualizados = 0
    detalle: list[dict] = []
    vistos: set[str] = set()

    for tercero in terceros:
        ident = _solo_digitos(tercero.get("identificacion", ""))
        if not ident:
            detalle.append({"identificacion": "", "nombre": tercero.get("nombre", ""),
                            "accion": "omitido"})
            continue

        if ident in indice:
            fila = indice[ident]
            accion = "actualizado"
            actualizados += 1
        else:
            ultima_fila += 1
            fila = ultima_fila
            indice[ident] = fila
            accion = "agregado"
            agregados += 1

        for campo, col in columnas.items():
            valor = tercero.get(campo, "")
            # No sobrescribir un valor existente con uno vacío.
            if valor in (None, "") and accion == "actualizado":
                continue
            ws.cell(row=fila, column=col, value=valor)

        vistos.add(ident)
        detalle.append({
            "identificacion": ident,
            "nombre": tercero.get("nombre", ""),
            "accion": accion,
        })

    buffer = io.BytesIO()
    wb.save(buffer)
    resumen = {
        "agregados": agregados,
        "actualizados": actualizados,
        "creado": creado,
        "columnas": sorted(columnas.keys()),
        "fila_encabezados": fila_enc,
        "detalle": detalle,
    }
    logger.info(
        "Maestro de terceros actualizado: %d agregados, %d actualizados (nuevo=%s).",
        agregados, actualizados, creado,
    )
    return buffer.getvalue(), resumen
