"""
Tests de la validación de tipo de archivo maestro (``app.maestros``).

Cubre el caso real que rompía el módulo de Terceros: subir el Plan de Cuentas en
la casilla de Terceros hacía que «el módulo usara el maestro de cuentas».
"""

import io

from openpyxl import Workbook

from app.config import FILA_ENCABEZADOS_MAESTROS
from app.maestros import (
    clasificar_maestro, clasificar_encabezados, validar_maestro,
)


def _xlsx(headers: list[str], fila_datos: list | None = None) -> bytes:
    """Crea un .xlsx con encabezados en la fila 7 (como los maestros reales)."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Empresa de prueba")
    for col, h in enumerate(headers, start=1):
        ws.cell(row=FILA_ENCABEZADOS_MAESTROS + 1, column=col, value=h)
    if fila_datos:
        for col, v in enumerate(fila_datos, start=1):
            ws.cell(row=FILA_ENCABEZADOS_MAESTROS + 2, column=col, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_HEADERS_TERCEROS = [
    "Nombre tercero", "Tipo de identificación", "Identificación",
    "Digito verificación", "Dirección", "Ciudad", "Teléfono", "Estado",
]
_HEADERS_CUENTAS = [
    "Código", "Nombre", "Nivel agrupación", "Naturaleza", "Activo",
]


class TestClasificar:
    def test_terceros(self):
        assert clasificar_maestro(_xlsx(_HEADERS_TERCEROS)) == "terceros"

    def test_cuentas(self):
        assert clasificar_maestro(_xlsx(_HEADERS_CUENTAS)) == "cuentas"

    def test_comprobantes(self):
        headers = ["Código", "Tipo de comprobante", "Nombre"]
        assert clasificar_maestro(_xlsx(headers)) == "comprobantes"

    def test_desconocido(self):
        assert clasificar_encabezados(["Col A", "Col B"]) == "desconocido"

    def test_nit_alias_se_reconoce_como_terceros(self):
        headers = ["Nombre", "NIT", "Dirección", "Ciudad"]
        assert clasificar_maestro(_xlsx(headers)) == "terceros"


class TestValidar:
    def test_terceros_correcto_no_error(self):
        assert validar_maestro("terceros", _xlsx(_HEADERS_TERCEROS)) is None

    def test_cuentas_en_casilla_terceros_da_error(self):
        msg = validar_maestro("terceros", _xlsx(_HEADERS_CUENTAS))
        assert msg is not None
        assert "Plan de Cuentas" in msg

    def test_terceros_en_casilla_cuentas_da_error(self):
        msg = validar_maestro("cuentas", _xlsx(_HEADERS_TERCEROS))
        assert msg is not None
        assert "Listado de Terceros" in msg

    def test_archivo_ilegible_no_bloquea(self):
        # Bytes que no son un Excel válido: la validación es permisiva.
        assert validar_maestro("terceros", b"esto no es un xlsx") is None

    def test_desconocido_no_bloquea(self):
        assert validar_maestro("terceros", _xlsx(["Col A", "Col B"])) is None


def test_actualizar_maestro_con_cuentas_da_error_claro():
    """Importar un RUT sobre un 'terceros' que en realidad es el plan de cuentas
    debe fallar con un mensaje que mencione el Plan de Cuentas."""
    import pytest
    from app.terceros_rut import actualizar_maestro_terceros, mapear_rut_a_tercero

    cuentas_bytes = _xlsx(_HEADERS_CUENTAS, fila_datos=["1105", "CAJA", "Transaccional", "Débito", "Sí"])
    tercero = mapear_rut_a_tercero({"nit": "8356245", "dv": "5", "nombre": "X"})

    with pytest.raises(ValueError) as exc:
        actualizar_maestro_terceros([tercero], cuentas_bytes)
    assert "Plan de Cuentas" in str(exc.value)
