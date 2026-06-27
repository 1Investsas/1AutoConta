"""
Tests del módulo de Terceros: lectura del RUT de la DIAN y actualización del
maestro de terceros.

- ``parsear_rut_words``: extracción posicional de los datos del RUT (persona
  jurídica y persona natural) a partir de las palabras con coordenadas.
- ``mapear_rut_a_tercero`` + ``actualizar_maestro_terceros``: mapeo a columnas
  del maestro y upsert (crear, agregar, actualizar) sobre el Excel, con
  verificación de ida y vuelta usando el lector real ``cargar_maestro_terceros``.
"""

import pytest

from app.rut import parsear_rut_words, RUTParseError
from app.terceros_rut import mapear_rut_a_tercero, actualizar_maestro_terceros


# ---------------------------------------------------------------------------
# Constructores de palabras (simulan la salida de pdfplumber.extract_words)
# ---------------------------------------------------------------------------

def _w(text: str, x0: float, top: float) -> dict:
    """Crea una palabra con su caja a partir de (texto, x inicial, fila)."""
    ancho = max(4.0, len(text) * 5.0)
    return {"text": text, "x0": x0, "x1": x0 + ancho, "top": top}


def _digitos(cadena: str, x_inicio: float, top: float, paso: float = 11.0) -> list[dict]:
    """Crea una palabra por cada dígito, separadas horizontalmente (como el RUT)."""
    return [_w(d, x_inicio + i * paso, top) for i, d in enumerate(cadena)]


def _words_juridica() -> list[dict]:
    """Palabras de la hoja principal de un RUT de persona jurídica."""
    words: list[dict] = []
    # Casilla 5/6: NIT 901331657 + DV 7 (fila 180)
    words += _digitos("901331657", 86, 180)
    words += [_w("7", 192, 180)]
    # Casilla 12: dirección seccional
    words += [_w("Impuestos", 205, 180), _w("de", 229, 180), _w("Medellín", 236, 180)]
    words += [_w("1", 455, 180), _w("1", 469, 180)]  # código seccional
    # Casilla 24: tipo de contribuyente (fila 216)
    words += [_w("Persona", 27, 216), _w("jurídica", 55, 216), _w("1", 173, 216)]
    # Casilla 35: razón social (fila 288)
    for i, t in enumerate(["1", "INVERSIONES", "ESTRATEGICAS", "SOCIEDAD", "POR",
                           "ACCIONES", "SIMPLIFICADA"]):
        words += [_w(t, 27 + i * 35, 288)]
    # Casillas 36/37: nombre comercial y sigla (fila 312)
    words += [_w("1", 27, 312), _w("INVEST", 34, 312), _w("S.A.S.", 65, 312)]
    words += [_w("INVEST", 346, 312), _w("S.A.S.", 377, 312)]
    # Casillas 38-40: ubicación (fila 348)
    words += [_w("COLOMBIA", 27, 348)] + _digitos("169", 168, 348, 10)
    words += [_w("Antioquia", 201, 348)] + _digitos("05", 371, 348, 10)
    words += [_w("Medellín", 393, 348)] + _digitos("001", 564, 348, 10)
    # Casilla 41: dirección principal (fila 373)
    for i, t in enumerate(["CR", "82", "A", "CL", "32", "A", "310"]):
        words += [_w(t, 27 + i * 18, 373)]
    # Casilla 42: correo (fila 384)
    words += [_w("contabilidad@1inversionesestrategicas.com", 106, 384)]
    # Casillas 44/45: teléfonos (fila 398) — con dígitos de etiqueta "1" y "2"
    words += [_w("1", 237, 398)] + _digitos("3193651539", 309, 398, 9)
    words += [_w("2", 434, 398)]
    return words


def _words_natural() -> list[dict]:
    """Palabras de la hoja principal de un RUT de persona natural."""
    words: list[dict] = []
    # Casilla 5/6: NIT 8356245 + DV 5 (fila 180)
    words += _digitos("8356245", 109, 180)
    words += [_w("5", 192, 180)]
    words += [_w("Impuestos", 205, 180), _w("de", 229, 180), _w("Medellín", 236, 180)]
    # Casilla 24/25/26 (fila 216)
    words += [_w("Persona", 27, 216), _w("natural", 55, 216), _w("o", 78, 216),
              _w("sucesión", 84, 216), _w("ilíquida", 114, 216), _w("2", 173, 216)]
    words += [_w("Cédula", 186, 216), _w("de", 210, 216), _w("Ciudadanía", 220, 216)]
    words += [_w("1", 287, 216), _w("3", 297, 216)]          # código tipo doc (13)
    words += _digitos("8356245", 337, 216, 11)                # número de identificación
    # Casillas 31-34: apellidos y nombres (fila 264)
    words += [_w("VERGARA", 27, 264), _w("CHICA", 155, 264),
              _w("JUAN", 282, 264), _w("CAMILO", 407, 264)]
    # Casillas 38-40: ubicación (fila 348)
    words += [_w("COLOMBIA", 27, 348)] + _digitos("169", 168, 348, 10)
    words += [_w("Antioquia", 201, 348)] + _digitos("05", 371, 348, 10)
    words += [_w("Medellín", 393, 348)] + _digitos("001", 564, 348, 10)
    # Casilla 41: dirección principal (fila 373)
    for i, t in enumerate(["CR", "82", "A", "33", "33", "AP", "202"]):
        words += [_w(t, 27 + i * 18, 373)]
    # Casilla 42: correo (fila 384)
    words += [_w("jcamver@gmail.com", 106, 384)]
    # Casillas 44/45: teléfonos (fila 398)
    words += [_w("1", 237, 398)] + _digitos("5598855", 337, 398, 9)
    words += [_w("2", 434, 398)] + _digitos("3193596428", 507, 398, 9)
    return words


# ---------------------------------------------------------------------------
# Parser del RUT — persona jurídica
# ---------------------------------------------------------------------------

class TestParserJuridica:
    @pytest.fixture
    def datos(self):
        return parsear_rut_words(
            _words_juridica(),
            texto_completo="48 - Impuesto sobre las ventas - IVA",
        )

    def test_nit_y_dv(self, datos):
        assert datos["nit"] == "901331657"
        assert datos["dv"] == "7"

    def test_tipo_persona(self, datos):
        assert datos["tipo_persona"] == "juridica"
        assert datos["tipo_identificacion"] == "NIT"

    def test_razon_social_y_sigla(self, datos):
        assert datos["razon_social"] == (
            "1 INVERSIONES ESTRATEGICAS SOCIEDAD POR ACCIONES SIMPLIFICADA"
        )
        assert datos["nombre"] == datos["razon_social"]
        assert datos["sigla"] == "INVEST S.A.S."

    def test_ubicacion(self, datos):
        assert datos["pais"] == "COLOMBIA"
        assert datos["departamento"] == "Antioquia"
        assert datos["ciudad"] == "Medellín"
        assert datos["direccion"] == "CR 82 A CL 32 A 310"

    def test_contacto(self, datos):
        assert datos["correo"] == "contabilidad@1inversionesestrategicas.com"
        # El dígito "1" de la etiqueta "Teléfono 1" no debe colarse en el número.
        assert datos["telefono1"] == "3193651539"
        assert datos["telefono2"] == ""

    def test_responsable_iva(self, datos):
        assert datos["responsable_iva"] is True
        assert datos["regimen_iva"] == "Responsable de IVA"


# ---------------------------------------------------------------------------
# Parser del RUT — persona natural
# ---------------------------------------------------------------------------

class TestParserNatural:
    @pytest.fixture
    def datos(self):
        return parsear_rut_words(_words_natural(), texto_completo="")

    def test_nit_y_dv(self, datos):
        assert datos["nit"] == "8356245"
        assert datos["dv"] == "5"

    def test_tipo_persona_y_documento(self, datos):
        assert datos["tipo_persona"] == "natural"
        assert datos["tipo_identificacion"] == "CC"

    def test_nombre_completo(self, datos):
        assert datos["primer_apellido"] == "VERGARA"
        assert datos["segundo_apellido"] == "CHICA"
        assert datos["primer_nombre"] == "JUAN"
        assert datos["otros_nombres"] == "CAMILO"
        # Nombre legible: nombres primero, luego apellidos.
        assert datos["nombre"] == "JUAN CAMILO VERGARA CHICA"
        assert datos["razon_social"] == ""

    def test_ubicacion_y_contacto(self, datos):
        assert datos["ciudad"] == "Medellín"
        assert datos["direccion"] == "CR 82 A 33 33 AP 202"
        assert datos["correo"] == "jcamver@gmail.com"
        assert datos["telefono1"] == "5598855"
        assert datos["telefono2"] == "3193596428"

    def test_no_responsable_iva(self, datos):
        assert datos["responsable_iva"] is False
        assert datos["regimen_iva"] == "No responsable de IVA"


def test_parser_sin_nit_lanza_error():
    """Un documento sin NIT reconocible no es un RUT válido."""
    with pytest.raises(RUTParseError):
        parsear_rut_words([_w("Hola", 100, 180)])


def test_cedula_de_10_digitos_no_pierde_el_primer_digito():
    """Una cédula de 10 dígitos arranca más a la izquierda (x≈75); el primer
    dígito no debe perderse."""
    words: list[dict] = []
    # NIT 1017189674 + DV 9 — el primer dígito en x=75 (cédula de 10 dígitos).
    words += _digitos("1017189674", 75, 180, 11)
    words += [_w("9", 193, 180)]
    # Identificación como persona natural + apellidos/nombres.
    words += [_w("Persona", 27, 216), _w("natural", 55, 216),
              _w("Cédula", 186, 216), _w("Ciudadanía", 220, 216)]
    words += [_w("HENAO", 27, 264), _w("YEPES", 155, 264),
              _w("ANA", 282, 264), _w("MARIA", 407, 264)]
    datos = parsear_rut_words(words)
    assert datos["nit"] == "1017189674"
    assert datos["dv"] == "9"
    assert datos["nombre"] == "ANA MARIA HENAO YEPES"


# ---------------------------------------------------------------------------
# Actualización del maestro de terceros
# ---------------------------------------------------------------------------

class TestActualizarMaestro:
    def _terceros(self):
        return [
            mapear_rut_a_tercero(parsear_rut_words(_words_juridica())),
            mapear_rut_a_tercero(parsear_rut_words(_words_natural())),
        ]

    def test_crea_maestro_nuevo(self):
        contenido, resumen = actualizar_maestro_terceros(self._terceros(), None)
        assert resumen["creado"] is True
        assert resumen["agregados"] == 2
        assert resumen["actualizados"] == 0
        assert contenido  # bytes no vacíos

    def test_round_trip_con_lector_real(self, tmp_path):
        """El archivo generado se lee con el lector real del maestro (fila 7)."""
        from app.importador import cargar_maestro_terceros

        contenido, _ = actualizar_maestro_terceros(self._terceros(), None)
        ruta = tmp_path / "Listado_de_Terceros.xlsx"
        ruta.write_bytes(contenido)

        df = cargar_maestro_terceros(str(ruta))
        assert len(df) == 2
        ids = set(df["Identificación"])
        assert ids == {"901331657", "8356245"}
        # La identificación queda normalizada (solo dígitos) y los nombres presentes.
        fila_pj = df[df["Identificación"] == "901331657"].iloc[0]
        assert "INVERSIONES ESTRATEGICAS" in fila_pj["Nombre tercero"]
        assert fila_pj["Tipo de identificación"] == "NIT"

    def test_upsert_actualiza_existente_y_agrega(self):
        # 1) Crear con un tercero.
        t_pj, t_pn = self._terceros()
        contenido, _ = actualizar_maestro_terceros([t_pj], None)

        # 2) Cambiar el teléfono del existente y agregar uno nuevo.
        t_pj_mod = dict(t_pj, telefono="6041112233")
        contenido2, resumen = actualizar_maestro_terceros([t_pj_mod, t_pn], contenido)

        assert resumen["agregados"] == 1       # el natural
        assert resumen["actualizados"] == 1    # el jurídico

        from app.importador import cargar_maestro_terceros
        import io as _io
        ruta = _io.BytesIO(contenido2)
        import openpyxl
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        # Solo deben existir dos filas de datos (no se duplicó el jurídico).
        ids = [
            ws.cell(row=r, column=3).value  # columna "Identificación"
            for r in range(8, ws.max_row + 1)
            if ws.cell(row=r, column=3).value
        ]
        assert ids.count("901331657") == 1
        assert "8356245" in [str(i) for i in ids]

    def test_no_sobrescribe_con_vacios(self):
        """Al actualizar, un valor vacío no debe borrar el dato existente."""
        t_pj = mapear_rut_a_tercero(parsear_rut_words(_words_juridica()))
        contenido, _ = actualizar_maestro_terceros([t_pj], None)

        t_vacio = dict(t_pj)
        t_vacio["telefono"] = ""   # sin teléfono nuevo
        contenido2, _ = actualizar_maestro_terceros([t_vacio], contenido)

        from app.importador import cargar_maestro_terceros
        import io as _io
        ruta = _io.BytesIO(contenido2)
        import openpyxl
        ws = openpyxl.load_workbook(ruta).active
        fila = next(r for r in range(8, ws.max_row + 1)
                    if ws.cell(row=r, column=3).value == "901331657")
        # El teléfono original sigue ahí (no fue borrado por el vacío).
        telefonos = [ws.cell(row=fila, column=c).value for c in range(1, ws.max_column + 1)]
        assert "3193651539" in [str(v) for v in telefonos]
